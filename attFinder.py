from datetime import datetime, timedelta
from supabase import create_client
import pandas as pd
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Supabase configuration
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')

# Initialize Supabase client
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
def calculate_aggregate_attendance():
    try:
        # Get all attendance records
        response = supabase.table('attendance').select('*').execute()
        attendance_data = response.data

        if not attendance_data:
            return {
                "success": False,
                "message": "No attendance records found"
            }

        # Process data by student
        student_attendance = {}
        dates_with_sessions = set()  # To track which dates had sessions

        for record in attendance_data:
            date = record['date']
            student_email = record['student_email']
            morning = record.get('morning_session')
            afternoon = record.get('afternoon_session')

            # Initialize student record if not exists
            if student_email not in student_attendance:
                student_attendance[student_email] = {
                    'present_sessions': 0,
                    'total_sessions': 0
                }

            # Track this date
            dates_with_sessions.add(date)

            # Count morning session if it exists
            if morning is not None:
                student_attendance[student_email]['total_sessions'] += 1
                if morning == 'present':
                    student_attendance[student_email]['present_sessions'] += 1

            # Count afternoon session if it exists
            if afternoon is not None:
                student_attendance[student_email]['total_sessions'] += 1
                if afternoon == 'present':
                    student_attendance[student_email]['present_sessions'] += 1

        # Calculate percentages and prepare report
        report = {
            'total_students': len(student_attendance),
            'total_dates': len(dates_with_sessions),
            'student_percentages': {},
            'class_average': 0,
            'attendance_summary': {
                '75_and_above': 0,
                'below_75': 0
            }
        }

        total_percentage = 0
        for student, data in student_attendance.items():
            if data['total_sessions'] > 0:
                percentage = (data['present_sessions'] / data['total_sessions']) * 100
                report['student_percentages'][student] = {
                    'percentage': round(percentage, 2),
                    'present_sessions': data['present_sessions'],
                    'total_sessions': data['total_sessions']
                }
                total_percentage += percentage

                # Count students above/below 75%
                if percentage >= 75:
                    report['attendance_summary']['75_and_above'] += 1
                else:
                    report['attendance_summary']['below_75'] += 1

        # Calculate class average
        if report['total_students'] > 0:
            report['class_average'] = round(total_percentage / report['total_students'], 2)

        return {
            "success": True,
            "data": report
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error calculating attendance: {str(e)}"
        }

def print_attendance_report():
    result = calculate_aggregate_attendance()

    if not result['success']:
        print(f"Error: {result['message']}")
        return

    data = result['data']

    print("\n=== ATTENDANCE REPORT ===")
    print(f"\nTotal Students: {data['total_students']}")
    print(f"Total Days with Sessions: {data['total_dates']}")
    print(f"Class Average Attendance: {data['class_average']}%")

    print("\n--- Attendance Distribution ---")
    print(f"Students with ≥75% attendance: {data['attendance_summary']['75_and_above']}")
    print(f"Students with <75% attendance: {data['attendance_summary']['below_75']}")

    print("\n--- Individual Student Reports ---")
    sorted_students = sorted(
        data['student_percentages'].items(),
        key=lambda x: x[1]['percentage'],
        reverse=True
    )

    for student, stats in sorted_students:
        status = "GOOD" if stats['percentage'] >= 75 else "NEEDS IMPROVEMENT"
        print(f"\nStudent: {student}")
        print(f"Attendance: {stats['percentage']}% ({stats['present_sessions']}/{stats['total_sessions']} sessions)")
        print(f"Status: {status}")

    print("\n=== End of Report ===\n")

# Usage:
# print_attendance_report()
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

def generate_detailed_attendance_excel():
    try:
        # Get raw attendance data
        response = supabase.table('attendance').select('*').order('date').execute()
        attendance_data = response.data

        if not attendance_data:
            return {"success": False, "message": "No attendance records found"}

        # Create workbook and styles
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        present_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        absent_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Organize data by student and date
        student_data = defaultdict(lambda: defaultdict(dict))
        all_dates = set()

        for record in attendance_data:
            date = record['date']
            student = record['student_email']
            all_dates.add(date)

            student_data[student][date] = {
                'morning': record.get('morning_session'),
                'afternoon': record.get('afternoon_session')
            }

        # Sort dates and students
        all_dates = sorted(list(all_dates))
        all_students = sorted(list(student_data.keys()))

        # Create Daily Attendance Sheet
        daily_sheet = wb.active
        daily_sheet.title = "Daily Attendance"

        # Headers
        headers = ["Student Email"]
        for date in all_dates:
            headers.extend([f"{date} Morning", f"{date} Afternoon"])
        headers.append("Overall %")

        for col, header in enumerate(headers, start=1):
            cell = daily_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Fill student attendance data
        for row_idx, student in enumerate(all_students, start=2):
            # Student email
            daily_sheet.cell(row=row_idx, column=1, value=student).border = border

            present_count = 0
            total_sessions = 0

            # Daily attendance
            col_idx = 2
            for date in all_dates:
                if date in student_data[student]:
                    # Morning session
                    morning = student_data[student][date]['morning']
                    if morning is not None:
                        cell = daily_sheet.cell(row=row_idx, column=col_idx)
                        cell.value = morning
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        if morning == 'present':
                            cell.fill = present_fill
                            present_count += 1
                        else:
                            cell.fill = absent_fill
                        total_sessions += 1
                    col_idx += 1

                    # Afternoon session
                    afternoon = student_data[student][date]['afternoon']
                    if afternoon is not None:
                        cell = daily_sheet.cell(row=row_idx, column=col_idx)
                        cell.value = afternoon
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        if afternoon == 'present':
                            cell.fill = present_fill
                            present_count += 1
                        else:
                            cell.fill = absent_fill
                        total_sessions += 1
                    col_idx += 1

            # Calculate and add overall percentage
            if total_sessions > 0:
                percentage = round((present_count / total_sessions) * 100, 2)
                cell = daily_sheet.cell(row=row_idx, column=col_idx)
                cell.value = f"{percentage}%"
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = present_fill if percentage >= 75 else absent_fill

        # Add summary row
        summary_row = daily_sheet.max_row + 2
        daily_sheet.cell(row=summary_row, column=1, value="Daily Average").font = Font(bold=True)

        # Calculate daily averages
        col_idx = 2
        for date in all_dates:
            for session in ['morning', 'afternoon']:
                present_count = sum(1 for student in all_students
                                  if date in student_data[student]
                                  and student_data[student][date][session] == 'present')
                total_count = sum(1 for student in all_students
                                if date in student_data[student]
                                and student_data[student][date][session] is not None)

                if total_count > 0:
                    percentage = round((present_count / total_count) * 100, 2)
                    cell = daily_sheet.cell(row=summary_row, column=col_idx)
                    cell.value = f"{percentage}%"
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = present_fill if percentage >= 75 else absent_fill
                col_idx += 1

        # Adjust column widths
        for column in daily_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            daily_sheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        filename = f"detailed_attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        return {
            "success": True,
            "message": f"Detailed Excel report generated successfully: {filename}"
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error generating Excel report: {str(e)}"
        }

# Usage:
result = generate_detailed_attendance_excel()
print(result['message'])